from __future__ import annotations

import codecs
import csv
import io
import json
import math
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import ijson
import pandas as pd
from openpyxl import load_workbook


class TabularImportError(ValueError):
    pass


@dataclass(frozen=True)
class FileRecordBatches:
    batches: Iterator[list[dict[str, Any]]]
    selected_sheet_name: str | None = None


def record_batches_from_file_path(
    file_path: str | Path,
    *,
    source_type: str,
    sheet_name: str | None = None,
    batch_size: int = 1000,
) -> FileRecordBatches:
    """Create a bounded-memory record stream for an uploaded tabular file."""
    if batch_size <= 0:
        raise ValueError("batch_size must be positive.")
    path = Path(file_path)
    if not path.is_file() or path.stat().st_size <= 0:
        raise TabularImportError("Uploaded file is empty.")

    selected_sheet_name = None
    if source_type == "csv":
        records = _iter_csv_records(path)
    elif source_type == "json":
        records = _iter_json_records(path)
    elif source_type == "txt":
        records = _iter_txt_records(path)
    elif source_type == "xlsx":
        selected_sheet_name = sheet_name
        if not selected_sheet_name:
            previews = xlsx_sheet_previews_from_path(path, preview_limit=1)
            if not previews.get("ok"):
                raise TabularImportError(str(previews.get("error") or "XLSX parsing failed."))
            sheets = list(previews.get("sheets") or [])
            selected = next(
                (item for item in sheets if item.get("selected")),
                sheets[0] if sheets else None,
            )
            if selected is None:
                raise TabularImportError("No importable XLSX sheet was found.")
            selected_sheet_name = str(selected["sheet_name"])
        records = _iter_xlsx_records(path, selected_sheet_name)
    else:
        raise TabularImportError(f"不支持的数据源类型: {source_type}")

    return FileRecordBatches(
        batches=_record_batches(records, batch_size=batch_size),
        selected_sheet_name=selected_sheet_name,
    )


def preview_file_from_path(
    file_path: str | Path,
    *,
    source_type: str,
    sheet_name: str | None = None,
    preview_limit: int = 8,
) -> dict[str, Any]:
    try:
        stream = record_batches_from_file_path(
            file_path,
            source_type=source_type,
            sheet_name=sheet_name,
        )
        preview: list[dict[str, Any]] = []
        row_count = 0
        columns: list[str] = []
        for batch in stream.batches:
            row_count += len(batch)
            if len(preview) < preview_limit:
                preview.extend(batch[: preview_limit - len(preview)])
            for record in batch:
                for column in record:
                    if column not in columns:
                        columns.append(column)
        if row_count == 0:
            return {"ok": False, "error": "文件中没有可导入记录。"}
        return {
            "ok": True,
            "row_count": row_count,
            "column_count": len(columns),
            "columns": columns,
            "preview_records": preview,
            "selected_sheet": stream.selected_sheet_name,
        }
    except (OSError, TabularImportError, ValueError) as exc:
        return {"ok": False, "error": str(exc)}


def xlsx_sheet_previews_from_path(
    file_path: str | Path,
    *,
    preview_limit: int = 8,
) -> dict[str, Any]:
    path = Path(file_path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"解析 XLSX sheet 失败: {exc}"}

    previews: list[dict[str, Any]] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            preview: list[dict[str, Any]] = []
            columns: list[str] = []
            row_count = 0
            for record in _iter_xlsx_worksheet_records(worksheet):
                row_count += 1
                for column in record:
                    if column not in columns:
                        columns.append(column)
                if len(preview) < preview_limit:
                    preview.append(record)
            if row_count == 0:
                continue
            previews.append(
                {
                    "sheet_name": str(sheet_name),
                    "row_count": row_count,
                    "column_count": len(columns),
                    "score": row_count * max(len(columns), 1),
                    "preview_records": preview,
                }
            )
    except Exception as exc:
        return {"ok": False, "error": f"解析 XLSX sheet 失败: {exc}"}
    finally:
        workbook.close()

    if not previews:
        return {"ok": False, "error": "未从 XLSX 中解析到有效 sheet。"}
    previews.sort(key=lambda item: int(item["score"]), reverse=True)
    for index, item in enumerate(previews):
        item["selected"] = index == 0
    return {"ok": True, "sheets": previews}


def records_from_file_bytes(
    *,
    file_bytes: bytes,
    source_type: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    if source_type == "csv":
        return {"ok": True, "data": _csv_records_from_bytes(file_bytes)}
    if source_type == "json":
        return _json_records_from_bytes(file_bytes)
    if source_type == "txt":
        return {"ok": True, "data": _txt_records_from_bytes(file_bytes)}
    if source_type != "xlsx":
        return {"ok": False, "error": f"不支持的数据源类型: {source_type}"}

    dataframe_result = xlsx_dataframe_from_bytes(file_bytes, sheet_name=sheet_name)
    if not dataframe_result["ok"]:
        return dataframe_result
    return {"ok": True, "data": dataframe_to_json_records(dataframe_result["data"])}


def xlsx_dataframe_from_bytes(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    selected_sheet_name = sheet_name
    try:
        sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
    except Exception as exc:
        return {"ok": False, "error": f"解析 XLSX 原始行数据失败: {exc}"}

    candidates = []
    for current_sheet_name, raw_df in sheets.items():
        normalized = _normalize_excel_sheet(raw_df)
        if normalized is None:
            continue
        score = int(normalized.shape[0] * max(normalized.shape[1], 1))
        candidates.append((score, str(current_sheet_name), normalized))

    if not candidates:
        return {
            "ok": False,
            "error": "未从 XLSX 中解析到有效表格数据，请检查 sheet、表头或空行。",
        }

    if selected_sheet_name:
        for _, candidate_name, normalized in candidates:
            if candidate_name == selected_sheet_name:
                return {"ok": True, "data": normalized}
        return {"ok": False, "error": f"未找到可导入的 sheet: {selected_sheet_name}"}

    candidates.sort(key=lambda item: item[0], reverse=True)
    return {"ok": True, "data": candidates[0][2]}


def xlsx_sheet_previews_from_bytes(
    file_bytes: bytes,
    *,
    preview_limit: int = 8,
) -> dict[str, Any]:
    try:
        sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
    except Exception as exc:
        return {"ok": False, "error": f"解析 XLSX sheet 失败: {exc}"}

    previews: list[dict[str, Any]] = []
    for sheet_name, raw_df in sheets.items():
        normalized = _normalize_excel_sheet(raw_df)
        if normalized is None:
            continue
        score = int(normalized.shape[0] * max(normalized.shape[1], 1))
        previews.append(
            {
                "sheet_name": str(sheet_name),
                "row_count": int(normalized.shape[0]),
                "column_count": int(normalized.shape[1]),
                "score": score,
                "preview_records": dataframe_to_json_records(normalized.head(preview_limit)),
            }
        )
    if not previews:
        return {"ok": False, "error": "未从 XLSX 中解析到有效 sheet。"}
    best_score = max(int(item["score"]) for item in previews)
    selected = False
    for item in previews:
        item["selected"] = not selected and int(item["score"]) == best_score
        selected = selected or bool(item["selected"])
    previews.sort(key=lambda item: int(item["score"]), reverse=True)
    return {"ok": True, "sheets": previews}


def dataframe_to_json_records(dataframe: Any) -> list[dict[str, Any]]:
    return json.loads(dataframe.to_json(orient="records", force_ascii=False, date_format="iso"))


def _csv_records_from_bytes(file_bytes: bytes) -> list[dict[str, str]]:
    text = _decode_text(file_bytes)
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _json_records_from_bytes(file_bytes: bytes) -> dict[str, Any]:
    text = _decode_text(file_bytes)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        return {"ok": False, "error": f"解析 JSON 失败: {exc}"}

    records = _records_from_json_payload(payload)
    if not records:
        return {"ok": False, "error": "JSON 中没有可导入的记录数组或对象。"}
    return {"ok": True, "data": records}


def _records_from_json_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [_json_row(item, index) for index, item in enumerate(payload, start=1)]
    if isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list):
                records = [_json_row(item, index) for index, item in enumerate(value, start=1)]
                if records:
                    return records
        return [_json_row(payload, 1)]
    return [{"value": payload}]


def _json_row(item: Any, index: int) -> dict[str, Any]:
    if isinstance(item, dict):
        return {str(key): value for key, value in item.items()}
    return {"row_number": index, "value": item}


def _txt_records_from_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    text = _decode_text(file_bytes)
    tabular = _delimited_text_records(text)
    if tabular:
        return tabular

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return [{"line_number": index, "text": line} for index, line in enumerate(lines, start=1)]


def _delimited_text_records(text: str) -> list[dict[str, str]]:
    delimiter = _detect_text_delimiter(text)
    if delimiter is None:
        return []
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    non_empty_rows = [[cell.strip() for cell in row] for row in rows if any(cell.strip() for cell in row)]
    if len(non_empty_rows) < 2 or max(len(row) for row in non_empty_rows) < 2:
        return []

    header = _dedupe_columns([cell or f"column_{index + 1}" for index, cell in enumerate(non_empty_rows[0])])
    records: list[dict[str, str]] = []
    for row in non_empty_rows[1:]:
        padded = row + [""] * max(len(header) - len(row), 0)
        records.append({header[index]: padded[index] for index in range(len(header))})
    return records


def _detect_text_delimiter(text: str) -> str | None:
    """Detect a table delimiter without relying only on csv.Sniffer.

    Sniffer can reject otherwise valid TSV files when the first data row contains
    very long list-valued cells or ASCII control separators (for example the
    public JDsearch samples).  A delimited header is stronger evidence and also
    avoids reading an arbitrary prefix that ends halfway through a large cell.
    """
    lines = [line for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        return None

    header = lines[0]
    candidates = ("\t", ",", ";", "|")
    header_ranked = sorted(candidates, key=header.count, reverse=True)
    for delimiter in header_ranked:
        column_count = header.count(delimiter) + 1
        if column_count < 2:
            continue
        matching_rows = sum(
            1
            for line in lines[1:11]
            if line.count(delimiter) + 1 == column_count
        )
        if matching_rows:
            return delimiter

    sample = "\n".join(lines[:10])[:32_768]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return None


def _decode_text(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _record_batches(
    records: Iterator[dict[str, Any]],
    *,
    batch_size: int,
) -> Iterator[list[dict[str, Any]]]:
    batch: list[dict[str, Any]] = []
    for record in records:
        batch.append(record)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


def _iter_csv_records(path: Path) -> Iterator[dict[str, Any]]:
    yield from _iter_delimited_file(path, delimiter=",", strip_cells=False)


def _iter_txt_records(path: Path) -> Iterator[dict[str, Any]]:
    encoding = _file_encoding(path)
    sample_lines: list[str] = []
    with path.open("r", encoding=encoding, errors="replace", newline="") as source:
        for line in source:
            if line.strip():
                sample_lines.append(line.rstrip("\r\n"))
            if len(sample_lines) >= 10:
                break
    delimiter = _detect_text_delimiter("\n".join(sample_lines))
    if delimiter is not None:
        yield from _iter_delimited_file(path, delimiter=delimiter, strip_cells=True)
        return

    with path.open("r", encoding=encoding, errors="replace") as source:
        line_number = 0
        for line in source:
            text = line.strip()
            if not text:
                continue
            line_number += 1
            yield {"line_number": line_number, "text": text}


def _iter_delimited_file(
    path: Path,
    *,
    delimiter: str,
    strip_cells: bool,
) -> Iterator[dict[str, Any]]:
    encoding = _file_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as source:
        rows = csv.reader(source, delimiter=delimiter)
        header: list[str] | None = None
        for row in rows:
            values = [cell.strip() for cell in row] if strip_cells else list(row)
            if not any(str(cell).strip() for cell in values):
                continue
            if header is None:
                header = _dedupe_columns(
                    [
                        str(cell).strip() or f"column_{index + 1}"
                        for index, cell in enumerate(values)
                    ]
                )
                continue
            padded = values + [""] * max(len(header) - len(values), 0)
            yield {
                header[index]: padded[index]
                for index in range(len(header))
            }


def _file_encoding(path: Path) -> str:
    with path.open("rb") as source:
        sample = source.read(65_536)
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            codecs.getincrementaldecoder(encoding)(errors="strict").decode(
                sample,
                final=False,
            )
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def _iter_json_records(path: Path) -> Iterator[dict[str, Any]]:
    array_prefix = _first_top_level_json_array(path)
    try:
        with path.open("rb") as source:
            if array_prefix is not None:
                items = ijson.items(source, f"{array_prefix}.item")
            elif _first_non_whitespace_byte(path) == b"[":
                items = ijson.items(source, "item")
            else:
                items = ijson.items(source, "", multiple_values=True)
            for index, item in enumerate(items, start=1):
                yield _json_row(_normalize_stream_json_value(item), index)
    except (ijson.JSONError, OSError, ValueError) as exc:
        raise TabularImportError(f"解析 JSON 失败: {exc}") from exc


def _first_top_level_json_array(path: Path) -> str | None:
    if _first_non_whitespace_byte(path) != b"{":
        return None
    try:
        with path.open("rb") as source:
            for prefix, event, _value in ijson.parse(source):
                if event == "start_array" and prefix and "." not in prefix:
                    return prefix
                if event == "end_map" and prefix == "":
                    return None
    except (ijson.JSONError, OSError, ValueError) as exc:
        raise TabularImportError(f"解析 JSON 失败: {exc}") from exc
    return None


def _first_non_whitespace_byte(path: Path) -> bytes:
    with path.open("rb") as source:
        while chunk := source.read(4096):
            for value in chunk:
                character = bytes((value,))
                if not character.isspace():
                    return character
    return b""


def _normalize_stream_json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, dict):
        return {
            str(key): _normalize_stream_json_value(item)
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [_normalize_stream_json_value(item) for item in value]
    return value


def _iter_xlsx_records(path: Path, sheet_name: str) -> Iterator[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise TabularImportError(f"解析 XLSX 原始行数据失败: {exc}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise TabularImportError(f"未找到可导入的 sheet: {sheet_name}")
        yield from _iter_xlsx_worksheet_records(workbook[sheet_name])
    finally:
        workbook.close()


def _iter_xlsx_worksheet_records(worksheet: Any) -> Iterator[dict[str, Any]]:
    sample_rows: list[tuple[Any, ...]] = []
    for row in _iter_non_empty_xlsx_rows(worksheet):
        sample_rows.append(row)
        if len(sample_rows) >= 20:
            break
    if not sample_rows:
        return

    header_index = _detect_excel_header_values(sample_rows)
    header_values = list(sample_rows[header_index])
    width = _meaningful_row_width(header_values)
    columns = _dedupe_columns(
        [
            str(value).strip() if not _is_blank_cell(value) else f"column_{index + 1}"
            for index, value in enumerate(header_values[:width])
        ]
    )

    for index, row in enumerate(_iter_non_empty_xlsx_rows(worksheet)):
        if index <= header_index:
            continue
        row_width = _meaningful_row_width(row)
        if row_width > len(columns):
            columns.extend(
                f"column_{column_index + 1}"
                for column_index in range(len(columns), row_width)
            )
        padded = list(row) + [None] * max(len(columns) - len(row), 0)
        record = {
            columns[column_index]: _json_compatible_cell(padded[column_index])
            for column_index in range(len(columns))
        }
        if any(not _is_blank_cell(value) for value in record.values()):
            yield record


def _iter_non_empty_xlsx_rows(worksheet: Any) -> Iterator[tuple[Any, ...]]:
    for row in worksheet.iter_rows(values_only=True):
        values = tuple(row)
        if any(not _is_blank_cell(value) for value in values):
            yield values


def _detect_excel_header_values(rows: list[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:20]):
        values = [value for value in row if not _is_blank_cell(value)]
        if not values:
            continue
        string_like = sum(isinstance(value, str) for value in values)
        unique_values = len({str(value).strip() for value in values})
        following_rows = max(len(rows) - index - 1, 0)
        score = unique_values * 3 + string_like * 2 + min(following_rows, 5)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _meaningful_row_width(row: list[Any] | tuple[Any, ...]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if not _is_blank_cell(row[index]):
            return index + 1
    return 0


def _json_compatible_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _normalize_excel_sheet(raw_df: Any) -> Any | None:
    if raw_df is None or raw_df.empty:
        return None
    df = raw_df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty:
        return None

    header_index = _detect_excel_header_row(df)
    header_values = [str(value).strip() for value in df.iloc[header_index].tolist()]
    data = df.iloc[header_index + 1 :].copy()
    data = data.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if data.empty:
        return None

    columns = _dedupe_columns(header_values[: len(data.columns)])
    if len(columns) < len(data.columns):
        columns.extend(f"column_{index + 1}" for index in range(len(columns), len(data.columns)))
    data.columns = columns[: len(data.columns)]
    data = data.loc[:, [column for column in data.columns if str(column).strip()]]
    data = data.where(pd.notna(data), None)
    return data if not data.empty else None


def _detect_excel_header_row(df: Any) -> int:
    best_index = 0
    best_score = -1
    max_scan = min(len(df), 20)
    for index in range(max_scan):
        row = df.iloc[index]
        values = [value for value in row.tolist() if not _is_blank_cell(value)]
        if not values:
            continue
        string_like = sum(1 for value in values if isinstance(value, str))
        unique_values = len({str(value).strip() for value in values})
        following_rows = max(len(df) - index - 1, 0)
        score = unique_values * 3 + string_like * 2 + min(following_rows, 5)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _dedupe_columns(values: list[str]) -> list[str]:
    columns: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values):
        base = value.strip() or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        columns.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return columns


def _is_blank_cell(value: Any) -> bool:
    return bool(pd.isna(value)) or str(value).strip() == ""
